from pathlib import Path
from typing import Iterable, Sequence, List, Dict, Any
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.common.config import get_settings
from app.common.storage import ensure_dir

settings = get_settings()


def write_rows_to_xlsx(headers: Sequence[str], rows: Iterable[Sequence], filename: str) -> Path:
    """Basic Excel export with headers and rows."""
    wb = Workbook()
    ws = wb.active
    ws.append(list(headers))
    for row in rows:
        ws.append(list(row))
    dest_dir = ensure_dir(settings.export_dir)
    path = dest_dir / filename
    wb.save(path)
    return path


def create_styled_excel(
    headers: List[str],
    rows: List[List[Any]],
    sheet_title: str = "Sheet1",
    bold_headers: bool = True,
    auto_width: bool = True,
    wrap_text_columns: List[int] = None,
) -> BytesIO:
    """
    Create a styled Excel workbook with formatting.
    
    Args:
        headers: Column headers
        rows: Data rows
        sheet_title: Title for the worksheet
        bold_headers: Whether to make headers bold
        auto_width: Whether to auto-adjust column widths
        wrap_text_columns: List of column indices (0-based) to enable text wrapping
    
    Returns:
        BytesIO object containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    
    # Add headers
    ws.append(headers)
    
    # Apply bold font to headers
    if bold_headers:
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
    
    # Add data rows
    for row in rows:
        ws.append(row)
    
    # Auto-adjust column widths
    if auto_width:
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Apply text wrapping to specified columns
    if wrap_text_columns:
        for col_idx in wrap_text_columns:
            col_letter = ws.cell(row=1, column=col_idx + 1).column_letter
            for cell in ws[col_letter]:
                cell.alignment = Alignment(wrap_text=True)
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file


def create_officials_excel(
    officials_data: List[Dict[str, Any]],
    filename: str = "unit_officials.xlsx",
) -> BytesIO:
    """
    Create Excel file for unit officials.
    
    Args:
        officials_data: List of official dictionaries
        filename: Name for the sheet
    
    Returns:
        BytesIO object containing the Excel file
    """
    headers = [
        "Unit Name",
        "President Name",
        "President Phone",
        "Vice President Name",
        "Vice President Phone",
        "Secretary Name",
        "Secretary Phone",
        "Joint Secretary Name",
        "Joint Secretary Phone",
        "Treasurer Name",
        "Treasurer Phone",
    ]
    
    rows = []
    for official in officials_data:
        rows.append([
            official.get("unit_name", ""),
            official.get("president_name", ""),
            official.get("president_phone", ""),
            official.get("vice_president_name", ""),
            official.get("vice_president_phone", ""),
            official.get("secretary_name", ""),
            official.get("secretary_phone", ""),
            official.get("joint_secretary_name", ""),
            official.get("joint_secretary_phone", ""),
            official.get("treasurer_name", ""),
            official.get("treasurer_phone", ""),
        ])
    
    return create_styled_excel(headers, rows, "Unit Officials")


def create_members_excel(
    members_data: List[Dict[str, Any]],
    filename: str = "unit_members.xlsx",
) -> BytesIO:
    """
    Create Excel file for unit members.
    
    Args:
        members_data: List of member dictionaries
        filename: Name for the sheet
    
    Returns:
        BytesIO object containing the Excel file
    """
    headers = [
        "Name",
        "Contact",
        "Age",
        "DoB",
        "Qualification",
        "Blood Group",
        "Unit Name",
    ]
    
    rows = []
    for member in members_data:
        rows.append([
            member.get("name", ""),
            f"+91 {member.get('number', '')}" if member.get('number') else "",
            member.get("age", ""),
            member.get("dob", ""),
            member.get("qualification", ""),
            member.get("blood_group", ""),
            member.get("unit_name", ""),
        ])
    
    return create_styled_excel(headers, rows, "Unit Members")


def create_councilors_excel(
    councilors_data: List[Dict[str, Any]],
    filename: str = "unit_councilors.xlsx",
) -> BytesIO:
    """
    Create Excel file for unit councilors.
    
    Args:
        councilors_data: List of councilor dictionaries
        filename: Name for the sheet
    
    Returns:
        BytesIO object containing the Excel file
    """
    headers = ["Name", "Contact", "Unit Name"]
    
    rows = []
    for councilor in councilors_data:
        rows.append([
            councilor.get("name", ""),
            f"+91 {councilor.get('number', '')}" if councilor.get('number') else "",
            councilor.get("unit_name", ""),
        ])
    
    return create_styled_excel(headers, rows, "Unit Councilors")


def create_conference_excel(
    district_info: Dict[str, Dict[str, Any]],
    conference_id: int,
) -> BytesIO:
    """
    Create Excel file for conference data aggregated by district.
    
    Args:
        district_info: Dictionary with district information
        conference_id: ID of the conference
    
    Returns:
        BytesIO object containing the Excel file
    """
    headers = ["District", "Type", "Unit", "Name", "Phone Number", "Gender", "Count"]
    rows = []
    
    for district, info in district_info.items():
        # Add officials
        for official in info.get('officials', []):
            rows.append([
                district,
                "Official",
                official.get('unit', ''),
                official.get('name', ''),
                official.get('phone', ''),
                official.get('gender', ''),
                '',
            ])
        
        # Add members
        for member in info.get('members', []):
            rows.append([
                district,
                "Member",
                member.get('unit', ''),
                member.get('name', ''),
                member.get('phone', ''),
                member.get('gender', ''),
                '',
            ])
        
        # Add counts row
        count_text = (
            f"Male(s): {info.get('count_of_total_male', 0)}\n"
            f"Female(s): {info.get('count_of_total_female', 0)}\n"
            f"Total: {info.get('total_count', 0)}\n"
            f"Veg: {info.get('veg_count', 0)}\n"
            f"Non-Veg: {info.get('non_veg_count', 0)}"
        )
        rows.append([
            district,
            "Counts",
            '',
            '',
            '',
            '',
            count_text,
        ])
    
    # Create Excel with text wrapping on Count column (index 6)
    return create_styled_excel(
        headers,
        rows,
        f"Conference Data",
        wrap_text_columns=[6]
    )


def create_payment_info_excel(
    district_info: Dict[str, Dict[str, Any]],
    conference_id: int,
) -> BytesIO:
    """
    Create Excel file for payment information aggregated by district.
    
    Args:
        district_info: Dictionary with district payment information
        conference_id: ID of the conference
    
    Returns:
        BytesIO object containing the Excel file
    """
    headers = ["District", "Type", "Name", "Phone", "Count of Members", "Count of Officials"]
    rows = []
    
    for district, info in district_info.items():
        # Add officials
        for official in info.get('officials', []):
            rows.append([
                district,
                "Official",
                official.get('name', ''),
                official.get('phone', ''),
                '',
                '',
            ])
        
        # Add members
        for member in info.get('members', []):
            rows.append([
                district,
                "Member",
                member.get('name', ''),
                member.get('phone', ''),
                '',
                '',
            ])
        
        # Add counts
        rows.append([
            district,
            "Counts",
            '',
            '',
            info.get('count_of_members', 0),
            info.get('count_of_officials', 0),
        ])
        
        # Add payments
        for payment in info.get('payments', []):
            rows.append([
                district,
                "Payment",
                str(payment.get('amount_to_pay', '')),
                payment.get('uploaded_by', ''),
                payment.get('date', ''),
                payment.get('status', ''),
            ])
    
    return create_styled_excel(headers, rows, "Payment Info")


# Kalamela-specific exports
def export_kalamela_call_sheet(
    individual_participations: Dict[str, List[Dict]],
    group_participations: Dict[str, Dict[str, List[Dict]]],
    district_name: str = "All Districts",
) -> BytesIO:
    """
    Create formatted Kalamela call sheet with:
    - Title with merged cells
    - Group events by event → by team (unit name)
    - Individual events by event
    - Columns: No., Chest Number, Code Number, Signature, Start Time, End Time, Remarks
    - Borders on all cells
    
    Args:
        individual_participations: Dict of event name → participants list
        group_participations: Dict of event name → team name → participants list
        district_name: Name of the district (for title)
    
    Returns:
        BytesIO object containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Call Sheet"
    
    # Title
    title = f"CSI Madhya Kerala Diocese Youth Movement Kalamela Call Sheet - {district_name}"
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_num = 3
    
    # Define borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["No.", "Chest Number", "Code Number", "Signature", "Start Time", "End Time", "Remarks"]
    
    # Group events first
    if group_participations:
        ws.merge_cells(f'A{row_num}:G{row_num}')
        section_cell = ws[f'A{row_num}']
        section_cell.value = "GROUP EVENTS"
        section_cell.font = Font(bold=True, size=12)
        section_cell.alignment = Alignment(horizontal='center')
        for col in range(1, 8):
            ws.cell(row=row_num, column=col).border = thin_border
        row_num += 1
        
        for event_name, teams in group_participations.items():
            # Event name
            ws.merge_cells(f'A{row_num}:G{row_num}')
            event_cell = ws[f'A{row_num}']
            event_cell.value = event_name
            event_cell.font = Font(bold=True)
            for col in range(1, 8):
                ws.cell(row=row_num, column=col).border = thin_border
            row_num += 1
            
            for team_name, participants in teams.items():
                # Team name
                ws.merge_cells(f'A{row_num}:G{row_num}')
                team_cell = ws[f'A{row_num}']
                team_cell.value = f"Team: {team_name}"
                team_cell.font = Font(italic=True)
                for col in range(1, 8):
                    ws.cell(row=row_num, column=col).border = thin_border
                row_num += 1
                
                # Headers for this team
                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=row_num, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.border = thin_border
                row_num += 1
                
                # Participants
                for idx, participant in enumerate(participants, start=1):
                    ws.cell(row=row_num, column=1, value=idx).border = thin_border
                    ws.cell(row=row_num, column=2, value=participant.get('participant_chest_number', '')).border = thin_border
                    ws.cell(row=row_num, column=3, value=participant.get('participant_id', '')).border = thin_border
                    ws.cell(row=row_num, column=4, value='').border = thin_border
                    ws.cell(row=row_num, column=5, value='').border = thin_border
                    ws.cell(row=row_num, column=6, value='').border = thin_border
                    ws.cell(row=row_num, column=7, value='').border = thin_border
                    row_num += 1
                
                row_num += 1  # Empty row between teams
    
    # Individual events
    if individual_participations:
        ws.merge_cells(f'A{row_num}:G{row_num}')
        section_cell = ws[f'A{row_num}']
        section_cell.value = "INDIVIDUAL EVENTS"
        section_cell.font = Font(bold=True, size=12)
        section_cell.alignment = Alignment(horizontal='center')
        for col in range(1, 8):
            ws.cell(row=row_num, column=col).border = thin_border
        row_num += 1
        
        for event_name, participants in individual_participations.items():
            # Event name
            ws.merge_cells(f'A{row_num}:G{row_num}')
            event_cell = ws[f'A{row_num}']
            event_cell.value = event_name
            event_cell.font = Font(bold=True)
            for col in range(1, 8):
                ws.cell(row=row_num, column=col).border = thin_border
            row_num += 1
            
            # Headers
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.border = thin_border
            row_num += 1
            
            # Participants
            for idx, participant in enumerate(participants, start=1):
                ws.cell(row=row_num, column=1, value=idx).border = thin_border
                ws.cell(row=row_num, column=2, value=participant.get('participant_chest_number', '')).border = thin_border
                ws.cell(row=row_num, column=3, value=participant.get('participant_id', '')).border = thin_border
                ws.cell(row=row_num, column=4, value='').border = thin_border
                ws.cell(row=row_num, column=5, value='').border = thin_border
                ws.cell(row=row_num, column=6, value='').border = thin_border
                ws.cell(row=row_num, column=7, value='').border = thin_border
                row_num += 1
            
            row_num += 1  # Empty row between events
    
    # Auto-adjust column widths
    for col_idx in range(1, 8):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file


def export_kalamela_chest_numbers(
    individual_participations: Dict[str, List[Dict]],
    district_name: str = "All Districts",
) -> BytesIO:
    """
    Export individual event chest numbers grouped by district.
    
    Columns: No., District, Chest Number, Participant, Event, Unit
    
    Args:
        individual_participations: Dict of event name → participants list
        district_name: Name of the district
    
    Returns:
        BytesIO object containing the Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Chest Numbers"
    
    # Title
    title = f"Individual Event Chest Numbers - {district_name}"
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_num = 3
    
    # Define borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["No.", "District", "Chest Number", "Participant", "Event", "Unit"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.border = thin_border
    row_num += 1
    
    # Flatten all participants
    all_participants = []
    for event_name, participants in individual_participations.items():
        for participant in participants:
            all_participants.append({
                **participant,
                "event_name": event_name,
            })
    
    # Sort by district, then chest number
    all_participants.sort(key=lambda x: (
        x.get('participant_district', ''),
        x.get('participant_chest_number', '')
    ))
    
    # Add data
    for idx, participant in enumerate(all_participants, start=1):
        ws.cell(row=row_num, column=1, value=idx).border = thin_border
        ws.cell(row=row_num, column=2, value=participant.get('participant_district', '')).border = thin_border
        ws.cell(row=row_num, column=3, value=participant.get('participant_chest_number', '')).border = thin_border
        ws.cell(row=row_num, column=4, value=participant.get('participant_name', '')).border = thin_border
        ws.cell(row=row_num, column=5, value=participant.get('event_name', '')).border = thin_border
        ws.cell(row=row_num, column=6, value=participant.get('participant_unit', '')).border = thin_border
        row_num += 1
    
    # Auto-adjust column widths
    for col_idx in range(1, 7):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file


def export_kalamela_results(
    individual_results: Dict[str, List[Dict]],
    group_results: Dict[str, List[Dict]],
) -> BytesIO:
    """
    Export all results for each event.
    
    Columns: No., Name, Unit, Event, Position, Points
    
    Args:
        individual_results: Dict of event name → all participants
        group_results: Dict of event name → all teams
    
    Returns:
        BytesIO object containing the Excel file
    """
    def format_position(position: int) -> str:
        """Format position number to ordinal (1st, 2nd, 3rd, 4th, etc.)"""
        if position == 0:
            return ""
        suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(
            position % 10 if position % 100 not in (11, 12, 13) else 0, 'th'
        )
        return f"{position}{suffix} Place"
    
    wb = Workbook()
    
    # Individual results sheet
    ws_ind = wb.active
    ws_ind.title = "Individual Results"
    
    # Title
    title = "Individual Event Results - All Results"
    ws_ind.merge_cells('A1:F1')
    title_cell = ws_ind['A1']
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_num = 3
    
    # Define borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["No.", "Name", "Unit", "Event", "Position", "Points"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_ind.cell(row=row_num, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.border = thin_border
    row_num += 1
    
    # Add individual results
    entry_num = 1
    for event_name, results in individual_results.items():
        for result in results:
            position = result.get('position', 0)
            position_text = format_position(position)
            
            ws_ind.cell(row=row_num, column=1, value=entry_num).border = thin_border
            ws_ind.cell(row=row_num, column=2, value=result.get('participant_name', '')).border = thin_border
            ws_ind.cell(row=row_num, column=3, value=result.get('unit_name', '')).border = thin_border
            ws_ind.cell(row=row_num, column=4, value=event_name).border = thin_border
            ws_ind.cell(row=row_num, column=5, value=position_text).border = thin_border
            ws_ind.cell(row=row_num, column=6, value=result.get('total_points', '')).border = thin_border
            row_num += 1
            entry_num += 1
    
    # Auto-adjust column widths
    for col_idx in range(1, 7):
        ws_ind.column_dimensions[get_column_letter(col_idx)].width = 20
    
    # Group results sheet
    ws_grp = wb.create_sheet("Group Results")
    
    # Title
    title = "Group Event Results - All Results"
    ws_grp.merge_cells('A1:E1')
    title_cell = ws_grp['A1']
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row_num = 3
    
    # Headers
    headers = ["No.", "Chest Number", "Event", "Position", "Points"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_grp.cell(row=row_num, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.border = thin_border
    row_num += 1
    
    # Add group results
    entry_num = 1
    for event_name, results in group_results.items():
        for result in results:
            position = result.get('position', 0)
            position_text = format_position(position)
            
            ws_grp.cell(row=row_num, column=1, value=entry_num).border = thin_border
            ws_grp.cell(row=row_num, column=2, value=result.get('chest_number', '')).border = thin_border
            ws_grp.cell(row=row_num, column=3, value=event_name).border = thin_border
            ws_grp.cell(row=row_num, column=4, value=position_text).border = thin_border
            ws_grp.cell(row=row_num, column=5, value=result.get('total_points', '')).border = thin_border
            row_num += 1
            entry_num += 1
    
    # Auto-adjust column widths
    for col_idx in range(1, 6):
        ws_grp.column_dimensions[get_column_letter(col_idx)].width = 20
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

