from django.db.models import Value, CharField
import re
from django.db.models import F, Q, Count, Min, Max

from .models import (
    IndividualEvent,
    GroupEvent,
    IndividualEventParticipation,
    GroupEventParticipation,
    KalamelaExcludeMembers,
    IndividualEventScoreCard,
)

from auth_app.models import UnitMembers, ClergyDistrict

from datetime import date


def view_all_individual_event_participants(request, district_id):
    if district_id:
        individual_event_participations = (
            IndividualEventParticipation.objects.filter(
                added_by__clergy_district_id=district_id
            )
            .select_related(
                "individual_event",
                "participant",
                "participant__registered_user",
                "participant__registered_user__unit_name",
                "participant__registered_user__unit_name__clergy_district",
            )
            .order_by("individual_event__name")
        )
    else:
        individual_event_participations = (
            IndividualEventParticipation.objects.all()
            .select_related(
                "individual_event",
                "participant",
                "participant__registered_user",
                "participant__registered_user__unit_name",
                "participant__registered_user__unit_name__clergy_district",
            )
            .order_by("individual_event__name")
        )

    individual_event_participations_custom_dict = {}

    for individual_event_particpant in individual_event_participations:

        event_name = individual_event_particpant.individual_event.name

        if event_name not in individual_event_participations_custom_dict:
            individual_event_participations_custom_dict[event_name] = []

        individual_event_participations_custom_dict[event_name].append(
            {
                "individual_event_participation_id": individual_event_particpant.id,
                "individual_event_id": individual_event_particpant.individual_event_id,
                "participant_id": individual_event_particpant.participant.id,
                "participant_name": individual_event_particpant.participant.name.title(),
                "participant_unit": individual_event_particpant.participant.registered_user.unit_name.name.title(),
                "participant_district": individual_event_particpant.participant.registered_user.unit_name.clergy_district.name.title(),
                "participant_phone": individual_event_particpant.participant.number,
            }
        )

    return individual_event_participations_custom_dict


def view_all_group_event_participants(request, district_id):
    if district_id:
        group_event_participations_obj = GroupEventParticipation.objects.filter(
            added_by__clergy_district_id=district_id
        )

    else:
        group_event_participations_obj = GroupEventParticipation.objects.all()
    group_event_participations = group_event_participations_obj.select_related(
        "group_event",
        "participant",
    ).order_by("group_event__name")

    group_event_participations_custom_dict = {}

    for grp_participant in group_event_participations:
        event_name = grp_participant.group_event.name
        team_code = grp_participant.participant.registered_user.unit_name.name.title()

        # Ensure event_name points to a dictionary
        if event_name not in group_event_participations_custom_dict:
            group_event_participations_custom_dict[event_name] = {}

        if team_code not in group_event_participations_custom_dict[event_name]:
            group_event_participations_custom_dict[event_name][team_code] = []

        # Append participant details
        group_event_participations_custom_dict[event_name][team_code].append(
            {
                "group_event_participation_id": grp_participant.id,
                "group_event_id": grp_participant.group_event_id,
                "group_event_max_allowed_limit": grp_participant.group_event.max_allowed_limit,
                "participant_unit_id": grp_participant.participant.registered_user.unit_name_id,
                "participant_id": grp_participant.participant_id,
                "participant_name": grp_participant.participant.name.title(),
                "participant_unit": grp_participant.participant.registered_user.unit_name.name.title(),
                "participant_district": grp_participant.participant.registered_user.unit_name.clergy_district.name.title(),
                "participant_phone": grp_participant.participant.number,
                "total_count": len(
                    group_event_participations_custom_dict[event_name][team_code]
                )
                + 1,
            }
        )
    return group_event_participations_custom_dict


from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from io import BytesIO
from django.http import HttpResponse

def admin_export_all_events_data(request, district_id, individual_event, group_event):
    group_event_participations_obj = GroupEventParticipation.objects.all()
    individual_event_participations_obj = IndividualEventParticipation.objects.all()

    # Filter by district
    district_name = None
    if district_id:
        district = ClergyDistrict.objects.get(id=district_id)
        district_name = district.name
        group_event_participations_obj = group_event_participations_obj.filter(
            added_by__clergy_district_id=district_id
        )
        individual_event_participations_obj = individual_event_participations_obj.filter(
            added_by__clergy_district_id=district_id
        )

    # Filter by individual and group events
    if individual_event:
        individual_event_participations_obj = individual_event_participations_obj.filter(
            individual_event_id=individual_event
        )

    if group_event:
        group_event_participations_obj = group_event_participations_obj.filter(
            group_event_id=group_event
        )

    # Process Group Event Participations
    group_event_participations = group_event_participations_obj.select_related(
        "group_event", "participant"
    ).order_by("group_event__name")

    group_event_participations_custom_dict = {}
    for grp_participant in group_event_participations:
        event_name = grp_participant.group_event.name
        team_code = grp_participant.participant.registered_user.unit_name.name.title()

        if event_name not in group_event_participations_custom_dict:
            group_event_participations_custom_dict[event_name] = {}

        if team_code not in group_event_participations_custom_dict[event_name]:
            group_event_participations_custom_dict[event_name][team_code] = []

        group_event_participations_custom_dict[event_name][team_code].append({
            "participant_name": grp_participant.participant.name.title(),
            "participant_chest_number": grp_participant.chest_number,
            "participant_unit": grp_participant.participant.registered_user.unit_name.name.title(),
            "participant_phone": grp_participant.participant.number,
        })

    # Process Individual Event Participations
    individual_event_participations = individual_event_participations_obj.select_related(
        "individual_event",
        "participant",
        "participant__registered_user",
        "participant__registered_user__unit_name",
        "participant__registered_user__unit_name__clergy_district",
    ).order_by("individual_event__name")

    individual_event_participations_custom_dict = {}
    for individual_participant in individual_event_participations:
        event_name = individual_participant.individual_event.name

        if event_name not in individual_event_participations_custom_dict:
            individual_event_participations_custom_dict[event_name] = []

        individual_event_participations_custom_dict[event_name].append({
            "participant_name": individual_participant.participant.name.title(),
            "participant_unit": individual_participant.participant.registered_user.unit_name.name.title(),
            "participant_chest_number": individual_participant.chest_number,
        })

    # Set up Excel file
    excel_title = f"Participants for {'All Districts' if not district_name else district_name}"
    excel_file_name = f"participants_{'all' if not district_name else district_name.lower()}"

    wb = Workbook()
    ws = wb.active
    ws.title = excel_file_name

    # Define styles
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=12)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Add title
    ws.merge_cells("A1:G1")
    ws["A1"] = "CSI Madhya Kerala Diocese"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align
    
    ws.merge_cells("A2:G2")
    ws["A2"] = "Youth Movement Kalamela"
    ws["A2"].font = title_font
    ws["A2"].alignment = center_align
    
    ws.merge_cells("A3:G3")
    ws["A3"] = "Call Sheet"
    ws["A3"].font = header_font
    ws["A3"].alignment = center_align
    
    ws.merge_cells("A4:G4")
    ws["A4"] = excel_title
    ws["A4"].font = title_font
    ws["A4"].alignment = center_align

    # Write Group Event Data
    current_row = 5
    if group_event:
        ws[f"A{current_row}"] = "Group Event Participants"
        ws[f"A{current_row}"].font = header_font
        current_row += 1

        for event_name, teams in group_event_participations_custom_dict.items():
            ws[f"A{current_row}"] = f"Item: {event_name}"
            ws[f"A{current_row}"].font = header_font
            current_row += 1

            for team, participants in teams.items():
                ws[f"A{current_row}"] = f"Team: {team}"
                current_row += 1

                # Add headers
                ws.append(["No.", "Chest Number", "Code Number", "Signature", "Start Time", "End Time", "Remarks"])
                for col in range(1, 8):
                    ws.cell(row=current_row, column=col).font = header_font
                    ws.cell(row=current_row, column=col).border = thin_border
                    ws.cell(row=current_row, column=col).alignment = center_align

                current_row += 1

                for idx, participant in enumerate(participants, start=1):
                    ws.append([
                        idx,
                        participant["participant_chest_number"],
                        "",
                        "",
                        "",
                        "",
                        ""
                    ])
                    for col in range(1, 8):
                        ws.cell(row=current_row, column=col).border = thin_border
                    current_row += 1
         # Save the workbook to an in-memory stream
        excel_stream = BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)

        # Create HTTP response
        response = HttpResponse(
            excel_stream,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="Participants_group_events.xlsx"'

        return response

    # Write Individual Event Data
    if individual_event:
        current_row += 2
        ws[f"A{current_row}"] = "Individual Event Participants"
        ws[f"A{current_row}"].font = header_font
        current_row += 1

        for event_name, participants in individual_event_participations_custom_dict.items():
            ws[f"A{current_row}"] = f"Event: {event_name}"
            ws[f"A{current_row}"].font = header_font
            current_row += 1

            # Add headers
            ws.append(["No.", "Chest Number", "Code Number", "Signature", "Start Time", "End Time", "Remarks"])
            for col in range(1, 8):
                ws.cell(row=current_row, column=col).font = header_font
                ws.cell(row=current_row, column=col).border = thin_border
                ws.cell(row=current_row, column=col).alignment = center_align

            current_row += 1

            for idx, participant in enumerate(participants, start=1):
                ws.append([
                        idx,
                        participant["participant_chest_number"],
                        "",
                        "",
                        "",
                        "",
                        ""
                    ])
                for col in range(1, 8):
                    ws.cell(row=current_row, column=col).border = thin_border
                current_row += 1

        # Save workbook
        # Create HTTP response
        excel_stream = BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)
        response = HttpResponse(
            excel_stream,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="participants_individual_events.xlsx"'

        return response

    if not individual_event and not group_event:
        ws[f"A{current_row}"] = "All Events"
        ws[f"A{current_row}"].font = header_font
        current_row += 1

        for event_name, teams in group_event_participations_custom_dict.items():
            ws[f"A{current_row}"] = f"Item: {event_name}"
            ws[f"A{current_row}"].font = header_font
            current_row += 1

            for team, participants in teams.items():
                ws[f"A{current_row}"] = f"Team: {team}"
                current_row += 1

                # Add headers
                ws.append(["No.", "Chest Number", "Code Number", "Signature", "Start Time", "End Time", "Remarks"])
                for col in range(1, 8):
                    ws.cell(row=current_row, column=col).font = header_font
                    ws.cell(row=current_row, column=col).border = thin_border
                    ws.cell(row=current_row, column=col).alignment = center_align

                current_row += 1

                for idx, participant in enumerate(participants, start=1):
                    ws.append([
                        idx,
                        participant["participant_chest_number"],
                        "",
                        "",
                        "",
                        "",
                        ""
                    ])
                    for col in range(1, 8):
                        ws.cell(row=current_row, column=col).border = thin_border
                    current_row += 1
        current_row += 2
        ws[f"A{current_row}"] = "Individual Event Participants"
        ws[f"A{current_row}"].font = header_font
        current_row += 1

        for event_name, participants in individual_event_participations_custom_dict.items():
            ws[f"A{current_row}"] = f"Event: {event_name}"
            ws[f"A{current_row}"].font = header_font
            current_row += 1

            # Add headers
            ws.append(["No.", "Chest Number", "Code Number", "Signature", "Start Time", "End Time", "Remarks"])
            for col in range(1, 8):
                ws.cell(row=current_row, column=col).font = header_font
                ws.cell(row=current_row, column=col).border = thin_border
                ws.cell(row=current_row, column=col).alignment = center_align

            current_row += 1

            for idx, participant in enumerate(participants, start=1):
                ws.append([
                        idx,
                        participant["participant_chest_number"],
                        "",
                        "",
                        "",
                        "",
                        ""
                    ])
                for col in range(1, 8):
                    ws.cell(row=current_row, column=col).border = thin_border
                current_row += 1

        # Save workbook
        # Create HTTP response
        excel_stream = BytesIO()
        wb.save(excel_stream)
        excel_stream.seek(0)
        response = HttpResponse(
            excel_stream,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="Participants_all__events.xlsx"'

        return response


def admin_export_all_chest_numbers(request):
    individual_event_participations_obj = IndividualEventParticipation.objects.all().order_by('added_by__clergy_district__name',
                                                                                              'chest_number')
    individual_event_participations_dict = {}
    
    for individual_participant in individual_event_participations_obj:
        district_name = individual_participant.participant.registered_user.unit_name.clergy_district.name

        if district_name not in individual_event_participations_dict:
            individual_event_participations_dict[district_name] = []

        individual_event_participations_dict[district_name].append({
            "participant_name": individual_participant.participant.name.title(),
            "participant_unit": individual_participant.participant.registered_user.unit_name.name.title(),
            "participant_chest_number": individual_participant.chest_number,
            "participant_event": individual_participant.individual_event.name,
            "participant_district": individual_participant.participant.registered_user.unit_name.clergy_district.name.title(),
        })
    
    # Set up Excel file
    excel_title = f"Chest Numbers for All Districts"
    excel_file_name = f"individual_events_chest_numbers"
    wb = Workbook()
    ws = wb.active
    ws.title = excel_file_name

    # Define styles
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=12)
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_text_align = Alignment(wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Add title
    ws.merge_cells("A1:G1")
    ws["A1"] = "CSI Madhya Kerala Diocese"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align
    
    ws.merge_cells("A2:G2")
    ws["A2"] = "Youth Movement Kalamela"
    ws["A2"].font = title_font
    ws["A2"].alignment = center_align
    
    ws.merge_cells("A3:G3")
    ws["A3"] = "Chest Numbers"
    ws["A3"].font = header_font
    ws["A3"].alignment = center_align
    
    ws.merge_cells("A4:G4")
    ws["A4"] = excel_title
    ws["A4"].font = title_font
    ws["A4"].alignment = center_align

    current_row = 5
    ws[f"A{current_row}"] = "Individual Events"
    ws[f"A{current_row}"].font = header_font
    current_row += 1

    # Loop through the districts and chest numbers
    for event_name, participants in individual_event_participations_dict.items():
            ws[f"A{current_row}"] = f"Event: {event_name}"
            ws[f"A{current_row}"].font = header_font
            current_row += 1

            # Add headers
            ws.append(["No.","District","Chest Number","Participant", "Event", "Unit"])
            for col in range(1, 7):
                ws.cell(row=current_row, column=col).font = header_font
                ws.cell(row=current_row, column=col).border = thin_border
                ws.cell(row=current_row, column=col).alignment = center_align

            current_row += 1

            for idx, participant in enumerate(participants, start=1):
                ws.append([
                        idx,
                        participant["participant_district"],
                        participant["participant_chest_number"],
                        participant["participant_name"],
                        participant["participant_event"],
                        participant["participant_unit"],
                    ])
                for col in range(1, 7):
                    ws.cell(row=current_row, column=col).border = thin_border
                current_row += 1

    # Save workbook to memory
    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)

    # Create HTTP response to download the Excel file
    response = HttpResponse(
        excel_stream,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{excel_file_name}.xlsx"'

    return response



def export_all_results(request):
    indivdiual_results_dict = {}
    # Fetch distinct individual event names
    individual_events = IndividualEvent.objects.all().values_list(
        'name',
        flat=True
    ).distinct()
    
    group_results_dict = {}
    group_events = GroupEvent.objects.all().values_list(
        'name',
        flat=True
    ).distinct()

    for event in individual_events:
        # Fetch top 3 results for the event
        
        individual_event_scores = IndividualEventScoreCard.objects.filter(
            event_participation__individual_event__name=event,
        ).exclude(added_on__isnull=True).select_related(
            'event_participation'
        ).order_by(
            'added_on',
            '-total_points'
        )[:3]  # Limit to the top 3 for each event
        
        if individual_event_scores.count() == 0:
            pass 
        else:
            indivdiual_results_dict[event] = individual_event_scores
    
    
    for event, individual_event_result in indivdiual_results_dict.items():
        for data in individual_event_result:
            print(data.participant)
            print(data.total_points)
    
    # Set up Excel file
    excel_title = f"Top 3 Results"
    excel_file_name = f"top_three_results"
    wb = Workbook()
    ws = wb.active
    ws.title = excel_file_name

    # Define styles
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=12)
    center_align = Alignment(horizontal="center", vertical="center")
    wrap_text_align = Alignment(wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Add title
    ws.merge_cells("A1:G1")
    ws["A1"] = "CSI Madhya Kerala Diocese"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align
    
    ws.merge_cells("A2:G2")
    ws["A2"] = "Youth Movement Kalamela"
    ws["A2"].font = title_font
    ws["A2"].alignment = center_align
    
    ws.merge_cells("A3:G3")
    ws["A3"] = "Chest Numbers"
    ws["A3"].font = header_font
    ws["A3"].alignment = center_align
    
    ws.merge_cells("A4:G4")
    ws["A4"] = excel_title
    ws["A4"].font = title_font
    ws["A4"].alignment = center_align

    current_row = 5
    ws[f"A{current_row}"] = "Individual Events"
    ws[f"A{current_row}"].font = header_font
    current_row += 1

    # Loop through the districts and chest numbers
    for event, individual_event_result in indivdiual_results_dict.items():
            ws[f"A{current_row}"] = f"Event: {event}"
            ws[f"A{current_row}"].font = header_font
            current_row += 1

            # Add headers
            ws.append(["No.","Name","Unit","Event","Postition"])
            for col in range(1, 6):
                ws.cell(row=current_row, column=col).font = header_font
                ws.cell(row=current_row, column=col).border = thin_border
                ws.cell(row=current_row, column=col).alignment = center_align

            current_row += 1

            for index, data in enumerate(individual_event_result):
                ws.append([
                        index,
                        data.participant.name,
                        data.participant.registered_user.unit_name.name,
                        data.event_participation.individual_event.name,
                        f"{index} Place"
                    ])
                for col in range(1, 7):
                    ws.cell(row=current_row, column=col).border = thin_border
                current_row += 1

    # Save workbook to memory
    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)

    # Create HTTP response to download the Excel file
    response = HttpResponse(
        excel_stream,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{excel_file_name}.xlsx"'

    return response





